from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

def generate_excel(data: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = data.get("title", "Sheet1")

    headers = data.get("headers", [])
    items = data.get("items", [])

    if headers:
        ws.append(headers)
        for cell in ws[1]:
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = __import__("openpyxl.styles").PatternFill(start_color="4F81BD", fill_type="solid")
            cell.alignment = Alignment(horizontal="center")

    if items:
        for item in items:
            ws.append(list(item.values()))
    else:
        ws.append(["데이터가 없습니다"])

    # 컬럼 너비 자동 조정
    for col in ws.columns:
        max_len = max((len(str(c.value or "")) for c in col), default=0)
        ws.column_dimensions[col[0].column_letter].width = (max_len + 2) * 1.2

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output.read()