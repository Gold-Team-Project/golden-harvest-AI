import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side


def generate_excel(data: dict):
    wb = Workbook()
    ws = wb.active
    ws.title = data.get("title", "Sheet1")

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4F81BD", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    headers = data.get("headers", [])
    ws.append(headers)

    for col_num, cell in enumerate(ws[1], 1):
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[chr(64 + col_num)].width = 15

    items = data.get("items", [])

    for item in items:
        row = []
        for h in headers:
            if h == "입고일자" or h == "출고일자":
                row.append(item.get("date", ""))
            elif h == "LOT번호":
                row.append(item.get("LOT", ""))
            elif h == "SKU번호":
                row.append(item.get("sku", ""))
            elif h == "수량":
                row.append(item.get("qty", 0))
            elif h == "단가":
                row.append(item.get("price", 0))
            elif h == "금액":
                row.append(item.get("amount", 0))
            else:
                row.append("")
        ws.append(row)

    for row in ws.iter_rows(min_row=2, max_row=len(items) + 1, min_col=1, max_col=len(headers)):
        for cell in row:
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output.read(), "xlsx", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
